import json
import glob
from inspect import currentframe
from icecream import ic
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

PLAYER_N = 3  # 三麻なら3
DORA_FANS = [31, 32, 33, 34]
RARE_FANS = [-1, 3, 4, 5, 6, 18, 19, 20, 24, 28]
ORIGIN_POINT = 35000


class HuleSingleData:
    def __init__(self, seat, isNagashi, rongPlayer, dadian, han, fu, fans):
        self.seat = seat
        self.isNagashi = isNagashi
        self.rongPlayer = rongPlayer  # ツモなら-1、ロンならseat
        self.dadian = dadian
        self.han = han
        self.fu = fu
        self.fans = fans[:]

    def getFansText(self):
        fanList = [(fan["id"], fan["val"]) for fan in self.fans]
        fanList.sort()
        return ",".join(
            [
                fanNames[fan[0]] + (str(fan[1]) if fan[0] in DORA_FANS else "")
                for fan in fanList
            ]
        )


class HandData:  # Hand: 局
    def __init__(self, resultDataJson):
        self.roundStr = ""
        self.deltaMain = [0] * PLAYER_N
        self.deltaSub = [0] * PLAYER_N
        self.huleData: list[HuleSingleData] = []
        CHANG = ["東", "南", "西"]
        JU = ["一", "二", "三", "四", "五", "六", "七", "八", "九", "十"]
        self.roundStr = (
            CHANG[resultDataJson["chang"]]
            + JU[resultDataJson["ju"]]
            + "局\n"
            + str(resultDataJson["ben"])
            + "本場"
        )

    def display(self):
        ic(self.roundStr)
        ic(self.deltaMain)
        ic(self.deltaSub)
        for hule in self.huleData:
            ic(hule.seat)
            ic(hule.isNagashi)
            ic(hule.rongPlayer)
            ic(hule.han)
            ic(hule.fu)
            fanList = [(fan["id"], fan["val"]) for fan in hule.fans]
            fanList.sort()
            print(
                *[
                    fanNames[fan[0]] + (str(fan[1]) if fan[0] in DORA_FANS else "")
                    for fan in fanList
                ],
                sep=",",
                end="",
            )
            print()


class RoundData:
    def __init__(self):
        self.uuid = ""
        self.names = [""] * PLAYER_N  # seat順
        self.scores = [0] * PLAYER_N
        self.thinkingTime = [0] * PLAYER_N
        self.stamps = [0] * PLAYER_N
        self.hands: list[HandData] = []

    def display(self):
        ic(self.names)
        for single in self.hands:
            single.display()


class PlayerHalfRoundData:
    def __init__(self):
        self.score = 0
        self.maxHule = 0
        self.paySum = 0
        self.doraCount = 0  # ドラの合計
        self.rareFans = []  # レア役のid
        self.thinkingTime = 0
        self.stamp = 0

    def reflectFans(self, fans):
        for fan in fans:
            if fan["id"] in DORA_FANS:
                self.doraCount += fan["val"]
            if fan["id"] in RARE_FANS:
                self.rareFans.append(fan["id"])


class PlayerData:
    def __init__(self):
        self.team = ""
        self.dataList: list[PlayerHalfRoundData] = []


fanNames: dict[int, str] = {}
playerDataDict: dict[str, PlayerData] = {}
roundDataList: list[RoundData] = []


def InitFanPairing():
    with open("fans.txt", "r") as f:
        while True:
            id = f.readline()[:-1]
            if not id:
                break
            fanname = f.readline()[:-1]
            fanNames[int(id)] = fanname


def InitPlayerData():
    with open("team.txt", "r") as f:
        while True:
            line = f.readline()
            if not line:
                break
            arr = line.split()
            nickname = arr[0]
            teamname = arr[1]
            playerDataDict[nickname] = PlayerData()
            playerDataDict[nickname].team = teamname


def CalcHalfRound(filename):
    with open(filename) as f:
        data = json.load(f)

    currentRoundData = RoundData()
    currentRoundData.uuid = data["head"]["uuid"]

    # head
    for player in data["head"]["result"]["players"]:
        currentRoundData.scores[player["seat"]] = player["total_point"]
    for account in data["head"]["accounts"]:
        # 第二回で追加
        if "seat" not in account:
            currentRoundData.names[0] = account["nickname"]
        else:
            currentRoundData.names[account["seat"]] = account["nickname"]

    ic(currentRoundData.names)

    # action
    currentScores = [0] * PLAYER_N
    for ai, action in enumerate(data["data"]["data"]["actions"]):
        if action["type"] == 1:
            # 局開始
            if action["result"]["name"] == ".lq.RecordNewRound":
                currentParent = action["result"]["data"]["ju"]
                currentRoundHandData = HandData(action["result"]["data"])
                currentScores = action["result"]["data"]["scores"]
                ic(currentScores)
            # 和了
            elif action["result"]["name"] == ".lq.RecordHule":
                for hule in action["result"]["data"][
                    "hules"
                ]:  # NOTE: ダブロンの可能性あり
                    if hule["zimo"]:  # ツモ
                        currentRoundHandData.huleData.append(
                            HuleSingleData(
                                seat=hule["seat"],
                                isNagashi=False,
                                rongPlayer=-1,
                                dadian=hule["dadian"],
                                han=hule["count"],
                                fu=hule["fu"],
                                fans=hule["fans"],
                            )
                        )
                        # それぞれの収支を計算
                        for seat in range(PLAYER_N):
                            if seat == hule["seat"]:
                                currentRoundHandData.deltaMain[seat] += hule["dadian"]
                            elif seat == currentParent:
                                currentRoundHandData.deltaMain[seat] -= hule[
                                    "point_zimo_qin"
                                ]
                            else:
                                currentRoundHandData.deltaMain[seat] -= hule[
                                    "point_zimo_xian"
                                ]
                    else:  # ロン
                        # 2つ前(状況によってはもう少し前)に放銃者の情報があるので、取ってくる
                        for back in range(2, 7):
                            if (
                                data["data"]["data"]["actions"][ai - back]["type"] != 1
                                or data["data"]["data"]["actions"][ai - back]["result"][
                                    "name"
                                ]
                                != ".lq.RecordDiscardTile"
                            ):
                                continue
                            rongPlayer = data["data"]["data"]["actions"][ai - back][
                                "result"
                            ]["data"]["seat"]
                            currentRoundHandData.huleData.append(
                                HuleSingleData(
                                    seat=hule["seat"],
                                    isNagashi=False,
                                    rongPlayer=rongPlayer,
                                    dadian=hule["dadian"],
                                    han=hule["count"],
                                    fu=hule["fu"],
                                    fans=hule["fans"],
                                )
                            )
                            currentRoundHandData.deltaMain[hule["seat"]] += hule[
                                "dadian"
                            ]
                            currentRoundHandData.deltaMain[rongPlayer] -= hule["dadian"]
                            break
                currentRoundData.hands.append(currentRoundHandData)

                for seat in range(PLAYER_N):
                    currentRoundHandData.deltaSub[seat] = (
                        (
                            action["result"]["data"]["old_scores"][seat]
                            - currentScores[seat]
                        )
                        + action["result"]["data"]["delta_scores"][seat]
                        - currentRoundHandData.deltaMain[seat]
                    )
            # 流局
            elif action["result"]["name"] == ".lq.RecordNoTile":
                for score in action["result"]["data"]["scores"]:
                    ic(action["result"]["data"]["scores"][0])
                    for seat in range(PLAYER_N):
                        currentRoundHandData.deltaSub[seat] = (
                            action["result"]["data"]["scores"][0]["old_scores"][seat]
                            - currentScores[seat]
                        ) + (
                            (
                                action["result"]["data"]["scores"][0]["delta_scores"][
                                    seat
                                ]
                            )
                            if "delta_scores" in action["result"]["data"]["scores"][0]
                            else 0
                        )
                    # 流し満貫
                    if "seat" in score:
                        huleSingleData = HuleSingleData(
                            seat=score["seat"],
                            isNagashi=True,
                            rongPlayer=-1,
                            han=0,
                            fu=0,
                            fans=[{"id": -1, "val": 0}],
                        )
                        currentRoundHandData.huleData.append(huleSingleData)
                currentRoundData.hands.append(currentRoundHandData)
        elif action["type"] == 2:
            if action["user_input"]["type"] == 1:  # スタンプ
                currentRoundData.stamps[action["user_input"]["seat"]] += 1
            elif action["user_input"]["type"] == 2:
                if "timeuse" in action["user_input"]["operation"]:
                    currentRoundData.thinkingTime[
                        action["user_input"]["seat"]
                    ] += action["user_input"]["operation"]["timeuse"]
            elif action["user_input"]["type"] == 3:
                if "timeuse" in action["user_input"]["cpg"]:
                    currentRoundData.thinkingTime[
                        action["user_input"]["seat"]
                    ] += action["user_input"]["cpg"]["timeuse"]

    roundDataList.append(currentRoundData)
    currentRoundData.display()


def CalcPlayerDataByRound(roundData: RoundData):
    records = [PlayerHalfRoundData() for _ in range(PLAYER_N)]
    for i in range(PLAYER_N):
        records[i].score = roundData.scores[i] / 1000
        records[i].thinkingTime = roundData.thinkingTime[i]
        records[i].stamp = roundData.stamps[i]

    for hand in roundData.hands:
        for i in range(PLAYER_N):
            records[i].maxHule = max(records[i].maxHule, hand.deltaMain[i])
            records[i].paySum += max(0, -hand.deltaMain[i])
        for hule in hand.huleData:
            for fan in hule.fans:
                if fan["id"] in DORA_FANS:
                    records[hule.seat].doraCount += fan["val"]
                if fan["id"] in RARE_FANS:
                    records[hule.seat].rareFans.append(fan["id"])

    for i in range(PLAYER_N):
        playerDataDict[roundData.names[i]].dataList.append(records[i])


def CalcPlayerData():  # roundDataからPlayerDataを計算
    for roundData in roundDataList:
        CalcPlayerDataByRound(roundData)


def Debug():
    for player, data in playerDataDict.items():
        ic(player)
        for playerRound in data.dataList:
            ic(playerRound.score)
            ic(playerRound.maxHule)
            ic(playerRound.paySum)
            ic(playerRound.doraCount)
            ic(playerRound.rareFans)


# TODO: 3人打ち対応


def ExportRoundSheet(ws, roundData: RoundData):
    for i in range(1, 100):
        for j in range(1, 100):
            ws.cell(row=i, column=j).font = Font(size=20)

    def DrawLineHorizontal(row, style):
        side = Side(style=style, color="000000")
        border = Border(bottom=side)
        for i in range(PLAYER_N + 3):
            ws[chr(ord("C") + i) + str(row)].border = border

    def FillHorizontal(row):
        for i in range(PLAYER_N + 3):
            ws[chr(ord("C") + i) + str(row)].fill = PatternFill(
                fgColor="EBF1DE", fill_type="solid"
            )

    ws["A1"] = "牌譜"
    ws["A1"].font = Font(color="0000FF", underline="single", size=20)
    ws["A1"].hyperlink = f"https://game.mahjongsoul.com/?paipu={roundData.uuid}"

    DIRECTIONS = ["東", "南", "西", "北"]
    for i in range(PLAYER_N):
        ws["{}2".format(chr(ord("D") + i))] = DIRECTIONS[i]
        ws["{}2".format(chr(ord("D") + i))].alignment = Alignment(
            horizontal="center", wrap_text=True
        )
        ws["{}2".format(chr(ord("D") + i))].font = Font(bold=True, size=20)

    DrawLineHorizontal(2, "thick")

    ws["C3"] = "雀魂HN"
    ws["C3"].alignment = Alignment(horizontal="center")

    DrawLineHorizontal(3, "thick")

    if PLAYER_N == 4:
        TEAM_COLOR = {
            "青": PatternFill(fgColor="C9DAF8", fill_type="solid"),
            "赤": PatternFill(fgColor="F4CCCC", fill_type="solid"),
            "白": PatternFill(fgColor="FFFFFF", fill_type="solid"),
            "黒": PatternFill(fgColor="D9D9D9", fill_type="solid"),
        }
    elif PLAYER_N == 3:
        TEAM_COLOR = {
            "A": PatternFill(fgColor="C9DAF8", fill_type="solid"),
            "B": PatternFill(fgColor="F4CCCC", fill_type="solid"),
            "C": PatternFill(fgColor="FFFFFF", fill_type="solid"),
        }

    for i in range(PLAYER_N):
        ws["{}3".format(chr(ord("D") + i))] = roundData.names[i]
        ws["{}3".format(chr(ord("D") + i))].fill = TEAM_COLOR[
            playerDataDict[roundData.names[i]].team
        ]
        ws["{}3".format(chr(ord("D") + i))].alignment = Alignment(
            horizontal="center", wrap_text=True
        )

    ws["{}3".format(chr(ord("H") + PLAYER_N - 4))] = "(供託)"
    ws["{}3".format(chr(ord("H") + PLAYER_N - 4))].alignment = Alignment(horizontal="center", wrap_text=True)
    ws["{}3".format(chr(ord("I") + PLAYER_N - 4))] = "(和了詳細)"
    ws["{}3".format(chr(ord("I") + PLAYER_N - 4))].alignment = Alignment(horizontal="center", wrap_text=True)

    row = 4
    scores = [ORIGIN_POINT] * PLAYER_N + [0]

    H = chr(ord("H") + PLAYER_N - 4)
    I = chr(ord("I") + PLAYER_N - 4)
    for hand in roundData.hands:
        FillHorizontal(row)

        DrawLineHorizontal(row, "thin")
        DrawLineHorizontal(row + 2, "thin")

        ws["C" + str(row + 1)] = hand.roundStr
        ws["C" + str(row + 1)].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.row_dimensions[row + 1].height = 23.5 * max(len(hand.huleData), 1)
        ws.row_dimensions[row + 2].height = 23.5 * max(len(hand.huleData), 1)
        ws.merge_cells(f"C{row + 1}:C{row + 2}")
        ws.merge_cells(f"{I}{row + 1}:{I}{row + 2}")
        ws[I + str(row + 1)].alignment = Alignment(wrap_text=True)

        huleText = ""
        for hule in hand.huleData:
            if len(huleText) > 0:
                huleText += "\n"
            if hule.isNagashi:
                huleText += "{} 流し満貫 8000".format(roundData.names[hule.seat])
            else:
                huleText += "{} {} {}\n{}".format(
                    roundData.names[hule.seat],
                    ("ツモ" if hule.rongPlayer == -1 else "ロン"),
                    hule.dadian,
                    hule.getFansText(),
                )
        ws[I + str(row + 1)] = huleText

        for i in range(PLAYER_N):
            ws[chr(ord("D") + i) + str(row)] = scores[i]
            ws[chr(ord("D") + i) + str(row)].number_format = "#,##0"
            if hand.deltaMain[i] != 0:
                ws[chr(ord("D") + i) + str(row + 1)] = hand.deltaMain[i]
                ws[chr(ord("D") + i) + str(row + 1)].number_format = "+#,##0;-#,###;0"
                ws[chr(ord("D") + i) + str(row + 1)].font = Font(
                    color=("0000FF" if hand.deltaMain[i] > 0 else "FF0000"), size=20
                )
            if hand.deltaSub[i] != 0:
                ws[chr(ord("D") + i) + str(row + 2)] = hand.deltaSub[i]
                ws[chr(ord("D") + i) + str(row + 2)].number_format = "+#,##0;-#,###;0"
                ws[chr(ord("D") + i) + str(row + 2)].font = Font(
                    color=("0000FF" if hand.deltaSub[i] > 0 else "FF0000"), size=20
                )
        ws[chr(ord("D") + PLAYER_N) + str(row)] = scores[PLAYER_N]
        ws[chr(ord("D") + PLAYER_N) + str(row)].number_format = "#,##0"

        deltaSubOther = -sum(hand.deltaSub)
        if deltaSubOther != 0:
            ws[chr(ord("D") + PLAYER_N) + str(row + 2)] = deltaSubOther
            ws[chr(ord("D") + PLAYER_N) + str(row + 2)].number_format = (
                "+#,##0;-#,###;0"
            )
            ws[chr(ord("D") + PLAYER_N) + str(row + 2)].font = Font(
                color=("0000FF" if hand.deltaSub[i] > 0 else "FF0000"), size=20
            )
        for i in range(PLAYER_N):
            scores[i] += hand.deltaMain[i] + hand.deltaSub[i]
        scores[PLAYER_N] += deltaSubOther
        row += 3
    for i in range(PLAYER_N + 1):
        ws[chr(ord("D") + i) + str(row)] = scores[i]
        ws[chr(ord("D") + i) + str(row)].number_format = "#,##0"
    FillHorizontal(row)
    DrawLineHorizontal(row, "thick")
    row += 1
    for i in range(PLAYER_N):
        ws[chr(ord("D") + i) + str(row)] = roundData.scores[i] / 1000
        ws[chr(ord("D") + i) + str(row)].number_format = "+#,##0.0;-#,##0.0;0.0"
        ws[chr(ord("D") + i) + str(row)].font = Font(
            color=("0000FF" if roundData.scores[i] > 0 else "FF0000"), size=20
        )

    ws.column_dimensions["C"].width = 15
    for i in range(PLAYER_N):
        ws.column_dimensions[chr(ord("D") + i)].width = 20
    ws.column_dimensions[H].width = 15
    ws.column_dimensions[I].width = 100

    def DrawLineVertical(col, style):
        side = Side(style=style, color="000000")
        for i in range(2, row + 1):
            top = ws[col + str(i)].border.top
            bottom = ws[col + str(i)].border.bottom
            right = ws[col + str(i)].border.right
            border = Border(left=side, right=right, top=top, bottom=bottom)
            ws[col + str(i)].border = border

    DrawLineVertical("D", "thick")
    DrawLineVertical(H, "medium")
    DrawLineVertical(I, "thick")


def ExportPlayerSheet(ws):
    for i in range(1, 100):
        for j in range(1, 100):
            ws.cell(row=i, column=j).font = Font(size=20)

    header = [
        "雀魂HN",
        "スコア",
        "最大和了",
        "支払合計",
        "ドラ合計",
        "レア役",
        "思考時間",
        "スタンプ",
    ]
    headerWidth = [25, 18, 18, 18, 15, 18, 18, 18]
    for i in range(len(header)):
        ws[chr(ord("B") + i) + str(2)] = header[i]
        ws[chr(ord("B") + i) + str(2)].font = Font(bold=True, size=20)
        ws[chr(ord("B") + i) + str(2)].alignment = Alignment(horizontal="center")
        ws.column_dimensions[chr(ord("B") + i)].width = headerWidth[i]

    row = 3
    for name, playerData in playerDataDict.items():
        ws["B" + str(row)] = name
        for playerSingleData in playerData.dataList:
            ws["C" + str(row)] = playerSingleData.score
            ws["C" + str(row)].number_format = "+#,##0.0;-#,##0.0;0.0"
            ws["D" + str(row)] = playerSingleData.maxHule
            ws["D" + str(row)].number_format = "#,##0"
            ws["E" + str(row)] = playerSingleData.paySum
            ws["E" + str(row)].number_format = "#,##0"
            ws["F" + str(row)] = playerSingleData.doraCount
            rareFansText = ",".join(
                [fanNames[fanId] for fanId in playerSingleData.rareFans]
            )
            ws["G" + str(row)] = rareFansText
            ws["H" + str(row)] = playerSingleData.thinkingTime
            ws["I" + str(row)] = playerSingleData.stamp
            row += 1


def ExportTotalResultSheet(ws):
    for i in range(1, 100):
        for j in range(1, 100):
            ws.cell(row=i, column=j).font = Font(size=20)

    for j in range(PLAYER_N):
        ws.column_dimensions[chr(ord("B") + j)].width = 18

    if PLAYER_N == 4:
        TEAM_COLOR = {
            "青": PatternFill(fgColor="C9DAF8", fill_type="solid"),
            "赤": PatternFill(fgColor="F4CCCC", fill_type="solid"),
            "白": PatternFill(fgColor="FFFFFF", fill_type="solid"),
            "黒": PatternFill(fgColor="D9D9D9", fill_type="solid"),
        }
    elif PLAYER_N == 3:
        TEAM_COLOR = {
            "A": PatternFill(fgColor="C9DAF8", fill_type="solid"),
            "B": PatternFill(fgColor="F4CCCC", fill_type="solid"),
            "C": PatternFill(fgColor="FFFFFF", fill_type="solid"),
        }

    if PLAYER_N == 4:
        teamToIndex = {"青": 0, "赤": 1, "白": 2, "黒": 3}
    elif PLAYER_N == 3:
        teamToIndex = {"A": 0, "B": 1, "C": 2}

    for team, color in TEAM_COLOR.items():
        if PLAYER_N == 4:
            ws[chr(ord("B") + teamToIndex[team]) + "2"] = team + "チーム"
        elif PLAYER_N == 3:
            ws[chr(ord("B") + teamToIndex[team]) + "2"] = "チーム" + team
        ws[chr(ord("B") + teamToIndex[team]) + "2"].fill = color

    row = 3
    for roundData in roundDataList:
        for i in range(PLAYER_N):
            col = teamToIndex[playerDataDict[roundData.names[i]].team]
            score = roundData.scores[i] / 1000
            ws[chr(ord("B") + col) + str(row)] = score
            ws[chr(ord("B") + col) + str(row)].number_format = "+#,##0.0;-#,##0.0;0.0"
            ws[chr(ord("B") + col) + str(row)].font = Font(
                color=("0000FF" if score > 0 else "FF0000"), size=20
            )
        row += 1


def ExportBook():
    wb = Workbook()
    wb.create_sheet("総合結果")
    ExportTotalResultSheet(wb["総合結果"])
    wb.create_sheet("プレイヤーデータ")
    ExportPlayerSheet(wb["プレイヤーデータ"])
    for i, roundData in enumerate(roundDataList):
        wb.create_sheet(f"試合{i+1}")
        ExportRoundSheet(wb[f"試合{i+1}"], roundData)
    wb.remove(wb["Sheet"])
    wb.save(f"result{PLAYER_N}.xlsx")


if __name__ == "__main__":
    InitPlayerData()
    InitFanPairing()

    for i in range(1, 7 + 1):
        file = f"paifu/{PLAYER_N}/{i}.json"
        ic(file)
        CalcHalfRound(file)

    CalcPlayerData()

    ExportBook()
    Debug()
